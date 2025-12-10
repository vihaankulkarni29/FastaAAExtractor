import pandas as pd
from pathlib import Path

genomes = ['CP107114.1', 'CP133923.1', 'CP133927.1', 'NZ_CP107117', 'NZ_CP107134', 'NZ_CP107172.1', 'NZ_CP107188.1']
target_genes = ['emrA', 'emrB', 'emrK', 'OXA-1', 'CTX-M-15', 'AAC(3)-IIe']

# Build presence matrix
data = []
for genome in genomes:
    row = {'Genome': genome}
    for gene in target_genes:
        # Check if file exists
        patterns = [
            f'results_5_target_genes/{genome}_{gene}.faa',
            f'results_5_target_genes/{genome.replace(".1", "")}_{gene}.faa'
        ]
        found = any(Path(p).exists() for p in patterns)
        row[gene] = 1 if found else 0
    row['Total'] = sum(row[g] for g in target_genes)
    data.append(row)

df = pd.DataFrame(data)
print('\n5 Target Genes Presence/Absence Matrix:\n')
print(df.to_string(index=False))

# Save to Excel with formatting
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

df.to_excel('Target_5_Genes_Matrix.xlsx', index=False, engine='openpyxl')

# Format the Excel
wb = load_workbook('Target_5_Genes_Matrix.xlsx')
ws = wb.active

# Header formatting
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Color code 1s and 0s
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=7):
    for cell in row:
        if cell.value == 1:
            cell.fill = green_fill
        elif cell.value == 0:
            cell.fill = red_fill
        cell.alignment = Alignment(horizontal='center')

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 3, 20)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save('Target_5_Genes_Matrix.xlsx')
print('\n✓ Saved to Target_5_Genes_Matrix.xlsx (color-coded)')
