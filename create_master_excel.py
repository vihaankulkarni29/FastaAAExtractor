import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

print('Collecting all extracted genes...')

# Define all genomes
genomes = ['NZ_CP107120', 'NZ_CP107134.1', 'NZ_CP107164', 'NZ_CP107178.1', 'NZ_HG941718']

# All result directories
result_dirs = {
    'CARD_Resistance': 'results_resistance_genes',
    'NCBI': 'results_ncbi',
    'ResFinder': 'results_resfinder',
    'Efflux_Pumps': 'results_efflux_pumps',
    'Efflux_Regulators': 'results_5genomes'
}

# Collect all genes
all_genes = {}

for db_name, result_dir in result_dirs.items():
    if os.path.exists(result_dir):
        files = list(Path(result_dir).glob('*.faa'))
        print(f'  {db_name}: {len(files)} files')
        
        for file in files:
            stem = file.stem
            genome = None
            gene = None
            
            # Match genome names
            for g in genomes:
                variants = [g, g.replace('.1', ''), g + '.1']
                for variant in variants:
                    if stem.startswith(variant + '_'):
                        genome = g
                        gene = stem[len(variant)+1:]
                        break
                if genome:
                    break
            
            if genome and gene:
                if gene not in all_genes:
                    all_genes[gene] = {'Database': db_name, 'Genomes': {g: 0 for g in genomes}}
                all_genes[gene]['Genomes'][genome] = 1

print(f'\nTotal unique genes collected: {len(all_genes)}')

# Create DataFrame
rows = []
for gene, data in all_genes.items():
    row = {'Gene': gene, 'Database': data['Database']}
    row.update(data['Genomes'])
    row['Total_Genomes'] = sum(data['Genomes'].values())
    rows.append(row)

df = pd.DataFrame(rows)

# Add detailed categorization
def categorize_gene(gene):
    gene_lower = gene.lower()
    
    # Beta-lactamases
    if 'ctx-m' in gene_lower:
        return 'Beta-lactamase (ESBL)', 'CTX-M ESBL - 3rd gen cephalosporin resistance'
    elif 'oxa' in gene_lower:
        return 'Beta-lactamase', 'OXA-type beta-lactamase'
    elif 'cmy' in gene_lower:
        return 'Beta-lactamase (AmpC)', 'CMY AmpC-type cephalosporinase'
    elif 'blaec' in gene_lower or gene.startswith('blaEC'):
        return 'Beta-lactamase (Chromosomal)', 'Chromosomal AmpC (intrinsic)'
    elif gene.startswith('bla'):
        return 'Beta-lactamase', 'Beta-lactamase enzyme'
    
    # Aminoglycoside resistance
    elif gene_lower.startswith('aac'):
        if 'cr' in gene_lower or 'ib' in gene_lower:
            return 'Aminoglycoside + Quinolone', 'AAC enzyme with quinolone resistance'
        return 'Aminoglycoside Resistance', 'Aminoglycoside acetyltransferase'
    
    # AcrAB family
    elif 'escherichia_coli_acra' in gene_lower or gene == 'acrA':
        return 'AcrAB-TolC Efflux', 'AcrA membrane fusion protein'
    elif gene == 'acrB':
        return 'AcrAB-TolC Efflux', 'AcrB RND transporter'
    elif gene in ['acrD', 'acrE', 'acrF']:
        return 'Acr Efflux Family', f'{gene.upper()} efflux transporter'
    elif gene == 'acrR':
        return 'Regulatory Protein', 'AcrR repressor of AcrAB'
    elif gene == 'acrS':
        return 'Regulatory Protein', 'AcrS regulator'
    
    # Emr efflux
    elif gene.startswith('emr'):
        comp = gene[3].upper()
        if comp == 'A':
            return 'EmrAB Efflux', 'EmrA membrane fusion protein'
        elif comp == 'B':
            return 'EmrAB Efflux', 'EmrB RND transporter'
        elif comp == 'K':
            return 'EmrKY Efflux', 'EmrK membrane fusion protein'
        elif comp == 'Y':
            return 'EmrKY Efflux', 'EmrY MFS transporter'
        elif comp == 'R':
            return 'Regulatory Protein', 'EmrR repressor'
        return 'Emr Efflux System', 'Emr efflux component'
    
    # Mdt efflux
    elif gene.startswith('mdt'):
        comp = gene[3].upper()
        if comp in ['A', 'B', 'C']:
            return 'MdtABC Efflux', f'MdtABC-TolC component {comp}'
        elif comp in ['E', 'F']:
            return 'MdtEF Efflux', f'MdtEF-TolC component {comp}'
        elif comp in ['N', 'O', 'P']:
            return 'MdtNOP Efflux', f'MdtNOP heteromultimer component {comp}'
        else:
            return 'Mdt Efflux System', f'Mdt{comp} MFS transporter'
    
    # MdfA
    elif 'mdfa' in gene_lower:
        return 'MdfA Efflux', 'MdfA multidrug MFS transporter'
    
    # TolC
    elif gene == 'tolC':
        return 'Outer Membrane Channel', 'TolC outer membrane efflux channel'
    
    # Regulatory proteins
    elif gene in ['marA', 'marR']:
        return 'Regulatory Protein (MarRAB)', f'{gene.upper()} - Mar regulon'
    elif gene in ['evgA', 'evgS']:
        return 'Regulatory Protein (EvgAS)', f'{gene.upper()} - EvgAS two-component'
    elif gene in ['cpxA', 'cpxR']:
        return 'Regulatory Protein (CpxAR)', f'{gene.upper()} - CpxAR two-component'
    elif gene == 'CRP':
        return 'Regulatory Protein', 'CRP - cAMP receptor protein'
    elif gene in ['gadW', 'gadX']:
        return 'Regulatory Protein', f'{gene.upper()} - Gad regulon'
    elif gene in ['baeR', 'baeS']:
        return 'Regulatory Protein (BaeRS)', f'{gene.upper()} - BaeRS two-component'
    elif gene in ['kdpE']:
        return 'Regulatory Protein', 'KdpE response regulator'
    elif gene == 'H-NS':
        return 'Nucleoid Protein', 'H-NS histone-like protein'
    
    # Other
    elif gene == 'yojI':
        return 'ABC Transporter', 'YojI - microcin J25 resistance'
    elif gene in ['bacA', 'ugd', 'pmrF', 'eptA']:
        return 'Cell Wall Modification', 'Lipid A or cell wall modification'
    elif gene == 'msbA':
        return 'Lipid Transporter', 'MsbA lipid flippase'
    else:
        return 'Other', 'Resistance or efflux-related protein'

categories = []
descriptions = []
for gene in df['Gene']:
    cat, desc = categorize_gene(gene)
    categories.append(cat)
    descriptions.append(desc)

df.insert(1, 'Category', categories)
df.insert(2, 'Function', descriptions)

# Sort
df = df.sort_values(['Category', 'Gene'])

# Create comprehensive summary
summary_data = []
for genome in genomes:
    total = df[genome].sum()
    beta_lac = df[df['Category'].str.contains('Beta-lactamase')][genome].sum()
    esbl = df[df['Category'] == 'Beta-lactamase (ESBL)'][genome].sum()
    carbapenem = df[df['Category'] == 'Beta-lactamase (Carbapenemase)'][genome].sum()
    amino = df[df['Category'].str.contains('Aminoglycoside')][genome].sum()
    efflux = df[df['Category'].str.contains('Efflux')][genome].sum()
    regulatory = df[df['Category'].str.contains('Regulatory')][genome].sum()
    
    # Determine resistance profile
    profile = []
    if esbl > 0:
        profile.append('ESBL-Producer')
    if df[df['Gene'].str.contains('CMY')][genome].sum() > 0:
        profile.append('AmpC-Producer')
    if carbapenem > 0:
        profile.append('Carbapenemase-Producer')
    if not profile:
        profile.append('Susceptible')
    
    summary_data.append({
        'Genome': genome,
        'Total_Genes': total,
        'Beta_lactamases': beta_lac,
        'ESBL': esbl,
        'Carbapenemase': carbapenem,
        'Aminoglycoside_Resistance': amino,
        'Efflux_Pumps': efflux,
        'Regulatory_Proteins': regulatory,
        'Resistance_Profile': ', '.join(profile)
    })

summary_df = pd.DataFrame(summary_data)

# Category breakdown
category_stats = df.groupby('Category').agg({
    'Gene': 'count',
    genomes[0]: 'sum',
    genomes[1]: 'sum',
    genomes[2]: 'sum',
    genomes[3]: 'sum',
    genomes[4]: 'sum'
})
category_stats.columns = ['Total_Genes'] + [f'{g}_Count' for g in genomes]
category_stats['Avg_Presence'] = category_stats[[f'{g}_Count' for g in genomes]].mean(axis=1).round(1)

# Save to Excel
print('\nCreating Excel file...')
with pd.ExcelWriter('MASTER_Gene_Presence_Absence.xlsx', engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name='Genome_Summary', index=False)
    df.to_excel(writer, sheet_name='Complete_Gene_Matrix', index=False)
    category_stats.to_excel(writer, sheet_name='Category_Breakdown')

# Format Excel
print('Formatting Excel file...')
wb = load_workbook('MASTER_Gene_Presence_Absence.xlsx')

# Format all sheets
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Special formatting for Complete_Gene_Matrix
ws = wb['Complete_Gene_Matrix']
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_font = Font(color='006100', bold=True)
red_font = Font(color='9C0006')

# Color code 1s and 0s
genome_cols = [4, 5, 6, 7, 8]  # Columns with genome data
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=genome_cols[0], max_col=genome_cols[-1]):
    for cell in row:
        if cell.value == 1:
            cell.fill = green_fill
            cell.font = green_font
        elif cell.value == 0:
            cell.fill = red_fill
            cell.font = red_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

wb.save('MASTER_Gene_Presence_Absence.xlsx')

print('\n✅ MASTER Excel file created successfully!')
print(f'   File: MASTER_Gene_Presence_Absence.xlsx')
print(f'   Total genes: {len(df)}')
print(f'   Genomes: {len(genomes)}')
print(f'\n📊 Sheets:')
print('   1. Genome_Summary - Overview of each genome')
print('   2. Complete_Gene_Matrix - Full presence/absence matrix (color-coded)')
print('   3. Category_Breakdown - Statistics by gene category')
print(f'\n🔬 Top gene categories:')
for cat, count in df['Category'].value_counts().head(10).items():
    print(f'   {cat}: {count}')
